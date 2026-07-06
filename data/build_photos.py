#!/usr/bin/env python3
"""
build_photos.py
ย่อรูปนักเรียนแล้ว copy ไปโฟลเดอร์ frontend/photos/
สร้าง students_photos.js ที่เก็บเฉพาะ path แทน base64

ผลลัพธ์:
    frontend/photos/{student_id}.jpg   — รูปย่อขนาด 200x200
    data/students_photos.js            — const STUDENT_PHOTOS = { "id": "photos/id.jpg" }
"""
import json, re, os
from pathlib import Path
from datetime import datetime

try:
    from PIL import Image
except ImportError:
    print("❌ ต้องติดตั้ง Pillow: pip3 install Pillow")
    exit(1)

BASE_DIR   = Path(__file__).parent.parent  # งานบันทึกความดี 69/
DATA_DIR   = Path(__file__).parent
PHOTOS_OUT = Path(__file__).parent.parent / "frontend" / "photos"
EXCEL_FILE = BASE_DIR / "รายชื่อ นพอ.ปี68 ทุกชั้นปี ( 5 ก.พ. 69 ตัดเกาหลี )SWD .xlsx"

# โฟลเดอร์รูปต้นฉบับ ในโฟลเดอร์จัดเก็บแบบมืออาชีพ data/raw_photos/
RAW_PHOTOS_DIR = DATA_DIR / "raw_photos"
PHOTO_DIRS = {
    68: RAW_PHOTOS_DIR / "นพอ.ชั้นปีที่ 1 ",
    67: RAW_PHOTOS_DIR / "ชั้นปีที่ 2",
    66: RAW_PHOTOS_DIR / "ชั้นปีที่ 3",
    65: RAW_PHOTOS_DIR / "ชั้นปีที่ 4",
}

SHEET_CLASS_MAP = {
    "SWD68": {"year": 2, "class_year": 68},
    "SWD67": {"year": 3, "class_year": 67},
    "SWD66": {"year": 4, "class_year": 66},
    "SWD65": {"year": 5, "class_year": 65},
    "SWD64": {"year": 5, "class_year": 64},
}

THUMB_SIZE = (200, 200)
JPEG_QUALITY = 75

# -------- helpers --------
def clean_id(val):
    if val is None: return None
    s = str(val).strip().replace("*", "").replace(" ", "")
    try: return str(int(float(s)))
    except: return None

def get_sorted_photos(folder: Path):
    if not folder.exists():
        print(f"  ⚠️  ไม่พบโฟลเดอร์: {folder}")
        return []
    imgs = []
    for f in folder.iterdir():
        if f.suffix.lower() in (".jpg", ".jpeg", ".png", ".gif", ".webp") and not f.name.startswith("."):
            m = re.match(r"^(\d+)", f.name.strip())
            num = int(m.group(1)) if m else 9999
            imgs.append((num, f))
    imgs.sort(key=lambda x: x[0])
    return imgs

def extract_name_from_filename(img_path: Path) -> str:
    stem = img_path.stem.strip()
    stem = re.sub(r"^\d+[_\s]+", "", stem)
    stem = re.sub(r"^นพอ\.\s*(\(ช\))?\s*", "", stem).strip()
    # ตัดข้อความท้ายที่ไม่ใช่ชื่อ เช่น "เลขที่ 34", "IMG_3197"
    stem = re.sub(r"\s+(เลขที่|IMG).*$", "", stem).strip()
    # ตัด _ ท้าย
    stem = stem.rstrip("_").strip()
    return stem

def resize_and_save(src_path: Path, dst_path: Path):
    """ย่อรูปเป็น 200x200 crop ตรงกลาง แล้วบันทึกเป็น JPEG"""
    img = Image.open(src_path)
    img = img.convert("RGB")

    # Center crop to square
    w, h = img.size
    side = min(w, h)
    left = (w - side) // 2
    top = (h - side) // 2
    img = img.crop((left, top, left + side, top + side))

    # Resize
    img = img.resize(THUMB_SIZE, Image.LANCZOS)
    img.save(dst_path, "JPEG", quality=JPEG_QUALITY, optimize=True)

# -------- read Excel --------
def read_students():
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    students = []
    
    # Load existing student profiles to merge edits (prevent data loss)
    existing_students = {}
    frontend_json_path = BASE_DIR / "frontend" / "data" / "students.json"
    if frontend_json_path.exists():
        try:
            with open(frontend_json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                existing_students = {s['student_id']: s for s in data}
            print(f"  Loaded {len(existing_students)} existing student records to preserve edits/passwords.")
        except Exception as e:
            print(f"  ⚠️ Failed to load existing students.json for merging: {e}")

    for sheet_name, meta in SHEET_CLASS_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠️  ไม่พบ sheet: {sheet_name}")
            continue
        ws = wb[sheet_name]
        row_num = 0
        for row in ws.iter_rows(min_row=5, values_only=True):
            sid = clean_id(row[1])
            if sid is None: continue
            rank  = str(row[2]).strip() if row[2] else "นพอ."
            fname = str(row[3]).strip() if row[3] else ""
            lname = str(row[4]).strip() if row[4] else ""
            row_num += 1
            
            password = sid
            email = ""
            telegram_chat_id = ""
            role = "student"
            class_year = meta["class_year"]
            year_level = meta["year"]
            note = ""
            
            # Merge edits from existing database if present
            if sid in existing_students:
                existing = existing_students[sid]
                rank = existing.get('rank', rank)
                fname = existing.get('first_name', fname)
                lname = existing.get('last_name', lname)
                password = existing.get('password', password)
                email = existing.get('email', email)
                telegram_chat_id = existing.get('telegram_chat_id', telegram_chat_id)
                role = existing.get('role', role)
                class_year = existing.get('class_year', class_year)
                year_level = existing.get('year_level', year_level)
                note = existing.get('note', note)

            students.append({
                "student_id": sid,
                "rank": rank,
                "first_name": fname,
                "last_name": lname,
                "full_name": f"{fname} {lname}".strip(),
                "class_year": class_year,
                "year_level": year_level,
                "password": password,
                "email": email,
                "telegram_chat_id": telegram_chat_id,
                "role": role,
                "note": note,
                "_row_num": row_num,
            })
    # Append Class 69 test students since they are not in the Excel sheets yet
    class69_students = [
        {
            "student_id": "6900001",
            "rank": "นพอ.",
            "first_name": "กิตติภพ",
            "last_name": "ทองดี",
            "full_name": "กิตติภพ ทองดี",
            "class_year": 69,
            "year_level": 1,
            "note": "นักเรียนใหม่ชั้นปีที่ 1",
            "password": "6900001",
            "email": "",
            "telegram_chat_id": "",
            "role": "student",
            "_row_num": 1
        },
        {
            "student_id": "6900002",
            "rank": "นพอ.",
            "first_name": "พรนภัส",
            "last_name": "จิตใจดี",
            "full_name": "พรนภัส จิตใจดี",
            "class_year": 69,
            "year_level": 1,
            "note": "นักเรียนใหม่ชั้นปีที่ 1",
            "password": "6900002",
            "email": "",
            "telegram_chat_id": "",
            "role": "student",
            "_row_num": 2
        },
        {
            "student_id": "6900003",
            "rank": "นพอ.",
            "first_name": "วรเมธ",
            "last_name": "รักสงบ",
            "full_name": "วรเมธ รักสงบ",
            "class_year": 69,
            "year_level": 1,
            "note": "นักเรียนใหม่ชั้นปีที่ 1",
            "password": "6900003",
            "email": "",
            "telegram_chat_id": "",
            "role": "student",
            "_row_num": 3
        },
        {
            "student_id": "6900004",
            "rank": "นพอ.",
            "first_name": "ชนม์นิภา",
            "last_name": "มีสุข",
            "full_name": "ชนม์นิภา มีสุข",
            "class_year": 69,
            "year_level": 1,
            "note": "นักเรียนใหม่ชั้นปีที่ 1",
            "password": "6900004",
            "email": "",
            "telegram_chat_id": "",
            "role": "student",
            "_row_num": 4
        },
        {
            "student_id": "6900005",
            "rank": "นพอ.",
            "first_name": "ปองพล",
            "last_name": "คนดี",
            "full_name": "ปองพล คนดี",
            "class_year": 69,
            "year_level": 1,
            "note": "นักเรียนใหม่ชั้นปีที่ 1",
            "password": "6900005",
            "email": "",
            "telegram_chat_id": "",
            "role": "student",
            "_row_num": 5
        }
    ]
    # Merge existing class69_students modifications if they exist in the DB
    for s in class69_students:
        sid = s['student_id']
        if sid in existing_students:
            existing = existing_students[sid]
            s['rank'] = existing.get('rank', s['rank'])
            s['first_name'] = existing.get('first_name', s['first_name'])
            s['last_name'] = existing.get('last_name', s['last_name'])
            s['full_name'] = existing.get('full_name', s['full_name'])
            s['class_year'] = existing.get('class_year', s['class_year'])
            s['year_level'] = existing.get('year_level', s['year_level'])
            s['password'] = existing.get('password', s['password'])
            s['email'] = existing.get('email', s['email'])
            s['telegram_chat_id'] = existing.get('telegram_chat_id', s['telegram_chat_id'])
            s['role'] = existing.get('role', s['role'])
            s['note'] = existing.get('note', s['note'])

    students.extend(class69_students)
    
    # Missing historical students from Main 2568.xlsx who are not in the new rosters
    missing_historical_students = [
        {
            "student_id": "6503719",
            "rank": "นพอ.",
            "first_name": "นันท์นภัส",
            "last_name": "เภสัชชา",
            "full_name": "นันท์นภัส เภสัชชา",
            "class_year": 65,
            "year_level": 5,
            "note": "ข้อมูลนำเข้าย้อนหลังจากประวัติปี 2568",
            "password": "6503719",
            "email": "",
            "telegram_chat_id": "",
            "role": "student",
            "_row_num": 991
        },
        {
            "student_id": "6603775",
            "rank": "นพอ.",
            "first_name": "ธิดารัตน์",
            "last_name": "นิลสังข์",
            "full_name": "ธิดารัตน์ นิลสังข์",
            "class_year": 66,
            "year_level": 4,
            "note": "ข้อมูลนำเข้าย้อนหลังจากประวัติปี 2568",
            "password": "6603775",
            "email": "",
            "telegram_chat_id": "",
            "role": "student",
            "_row_num": 992
        },
        {
            "student_id": "6703840",
            "rank": "นพอ.",
            "first_name": "ดลภัค",
            "last_name": "แก้วเอก",
            "full_name": "ดลภัค แก้วเอก",
            "class_year": 67,
            "year_level": 3,
            "note": "ข้อมูลนำเข้าย้อนหลังจากประวัติปี 2568",
            "password": "6703840",
            "email": "",
            "telegram_chat_id": "",
            "role": "student",
            "_row_num": 993
        },
        {
            "student_id": "6703850",
            "rank": "นพอ.",
            "first_name": "ประริชญา",
            "last_name": "ประสิทธิ์พรม",
            "full_name": "ประริชญา ประสิทธิ์พรม",
            "class_year": 67,
            "year_level": 3,
            "note": "ข้อมูลนำเข้าย้อนหลังจากประวัติปี 2568",
            "password": "6703850",
            "email": "",
            "telegram_chat_id": "",
            "role": "student",
            "_row_num": 994
        },
        {
            "student_id": "6703871",
            "rank": "นพอ.",
            "first_name": "วีรภัทร",
            "last_name": "นกดำ",
            "full_name": "วีรภัทร นกดำ",
            "class_year": 67,
            "year_level": 3,
            "note": "ข้อมูลนำเข้าย้อนหลังจากประวัติปี 2568",
            "password": "6703871",
            "email": "",
            "telegram_chat_id": "",
            "role": "student",
            "_row_num": 995
        }
    ]
    
    # Merge existing modifications for missing historical students
    for s in missing_historical_students:
        sid = s['student_id']
        if sid in existing_students:
            existing = existing_students[sid]
            s['rank'] = existing.get('rank', s['rank'])
            s['first_name'] = existing.get('first_name', s['first_name'])
            s['last_name'] = existing.get('last_name', s['last_name'])
            s['full_name'] = existing.get('full_name', s['full_name'])
            s['class_year'] = existing.get('class_year', s['class_year'])
            s['year_level'] = existing.get('year_level', s['year_level'])
            s['password'] = existing.get('password', s['password'])
            s['email'] = existing.get('email', s['email'])
            s['telegram_chat_id'] = existing.get('telegram_chat_id', s['telegram_chat_id'])
            s['role'] = existing.get('role', s['role'])
            s['note'] = existing.get('note', s['note'])

    students.extend(missing_historical_students)
    students.sort(key=lambda x: (x["class_year"], x["student_id"]))
    return students

# -------- match and resize --------
def build_photos(students):
    PHOTOS_OUT.mkdir(parents=True, exist_ok=True)
    photo_map = {}  # student_id → "photos/{id}.jpg"
    total_matched = 0

    for class_year, folder in PHOTO_DIRS.items():
        sorted_imgs = get_sorted_photos(folder)
        if not sorted_imgs:
            continue

        class_students = sorted(
            [s for s in students if s["class_year"] == class_year],
            key=lambda x: x["_row_num"]
        )
        name_to_stu = {f'{s["first_name"]} {s["last_name"]}'.strip(): s
                       for s in class_students}

        matched = 0
        for img_num, img_path in sorted_imgs:
            stu = None
            name_in_file = extract_name_from_filename(img_path)

            # 1) จับคู่จากชื่อในชื่อไฟล์
            stu = name_to_stu.get(name_in_file)

            # 2) จับคู่จากหมายเลข
            if stu is None:
                stu = next((s for s in class_students if s["_row_num"] == img_num), None)

            if stu:
                dst = PHOTOS_OUT / f"{stu['student_id']}.jpg"
                try:
                    resize_and_save(img_path, dst)
                    photo_map[stu["student_id"]] = f"photos/{stu['student_id']}.jpg"
                    matched += 1
                except Exception as e:
                    print(f"    ⚠️  {img_path.name}: {e}")

        print(f"  📷 รุ่น {class_year}: จับคู่ {matched}/{len(class_students)} คน (มีรูป {len(sorted_imgs)} ไฟล์)")
        total_matched += matched

    # Assign default photo for anyone who wasn't matched
    for stu in students:
        if stu["student_id"] not in photo_map:
            photo_map[stu["student_id"]] = "photos/chibi/chibi_lv1.png"
            total_matched += 1

    return photo_map, total_matched

# -------- write output --------
def write_photos_js(photo_map):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    js = (
        f"// Auto-generated student photos — {ts}\n"
        f"// นักเรียนที่มีรูป: {len(photo_map)} คน  DO NOT EDIT MANUALLY\n"
        f"// รูปถูกเก็บเป็นไฟล์ JPEG ใน frontend/photos/\n"
        f"const STUDENT_PHOTOS = {json.dumps(photo_map, ensure_ascii=False, indent=2)};\n"
    )
    output = DATA_DIR / "students_photos.js"
    output.write_text(js, encoding="utf-8")
    
    # เขียน JSON file
    json_path = DATA_DIR / "photos.json"
    json_path.write_text(json.dumps(photo_map, ensure_ascii=False, indent=2), encoding="utf-8")
    
    # Write to frontend/data/
    frontend_output = BASE_DIR / "frontend" / "data" / "students_photos.js"
    frontend_json_output = BASE_DIR / "frontend" / "data" / "photos.json"
    frontend_output.parent.mkdir(parents=True, exist_ok=True)
    frontend_output.write_text(js, encoding="utf-8")
    frontend_json_output.write_text(json.dumps(photo_map, ensure_ascii=False, indent=2), encoding="utf-8")

    size_kb = output.stat().st_size / 1024
    print(f"\n✅ students_photos.js ({size_kb:.0f} KB, เฉพาะ path) และ photos.json")
    print(f"   Sync-copied to {frontend_output} และ photos.json")
    return output

def write_students_js(students, photo_ids):
    """เขียน students_data.js ทั้งหมด (นักเรียนทุกคน)"""
    clean = [{k: v for k, v in s.items() if not k.startswith("_")} for s in students]
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    js = (
        f"// Auto-generated student data - DO NOT EDIT MANUALLY\n"
        f"// Generated from: รายชื่อ นพอ.ปี69 ทุกชั้นปี\n"
        f"// นักเรียนทั้งหมด: {len(clean)} คน\n\n"
        "const STUDENTS_DATA = "
        + json.dumps(clean, ensure_ascii=False, indent=2) + ";\n"
    )
    output = DATA_DIR / "students_data.js"
    output.write_text(js, encoding="utf-8")

    # Write to frontend/data/
    frontend_output = BASE_DIR / "frontend" / "data" / "students_data.js"
    frontend_output.write_text(js, encoding="utf-8")

    print(f"✅ students_data.js ({len(clean)} คน)")
    print(f"   Sync-copied to {frontend_output}")

    from collections import Counter
    years = Counter(s["class_year"] for s in students)
    labels = {68:"ชั้นปีที่ 2", 67:"ชั้นปีที่ 3", 66:"ชั้นปีที่ 4", 65:"ศิษย์เก่า 65"}
    for cy in sorted(years, reverse=True):
        print(f"   รุ่น {cy} ({labels.get(cy,'?')}): {years[cy]} คน")


# -------- main --------
def main():
    print("📂 อ่านข้อมูลนักเรียนจาก Excel...")
    students = read_students()
    print(f"   นักเรียนทั้งหมด: {len(students)} คน")

    print("\n🖼️  ย่อรูปและ copy ไป frontend/photos/...")
    photo_map, total = build_photos(students)

    write_photos_js(photo_map)
    write_students_js(students, set(photo_map.keys()))

    # ขนาดโฟลเดอร์รูป
    total_size = sum(f.stat().st_size for f in PHOTOS_OUT.iterdir() if f.is_file())
    print(f"\n📁 frontend/photos/ — {total} รูป ({total_size/1024/1024:.1f} MB)")
    print("🎉 เสร็จสิ้น! รูปพร้อมใช้งานผ่าน URL ไม่ต้อง base64")

if __name__ == "__main__":
    main()
