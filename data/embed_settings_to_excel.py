#!/usr/bin/env python3
"""Write NON-SECRET operational settings only.

Security rule:
- Never write passwords, bot tokens, API keys, chat IDs, LINE IDs, or personal data
  into Excel settings files.
- Authentication and notification credentials must live server-side only
  (Apps Script Properties / deployment secrets / environment variables).
"""
import os
import openpyxl

DATA_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(DATA_DIR)

EXCEL_FILES = [
    os.path.join(DATA_DIR, "รายชื่อ_นพอ_ปี69_2569.xlsx"),
    os.path.join(DATA_DIR, "รายชื่อ_นพอ_อัปเดตล่าสุด_2569.xlsx"),
    os.path.join(BASE_DIR, "รายชื่อ นพอ.ปี 69.xlsx"),
]

# Public/non-secret configuration only.
SETTINGS_DATA = [
    ["Setting_Key", "Setting_Value", "Description"],
    ["min_hours_per_semester", "25", "เกณฑ์ชั่วโมงขั้นต่ำต่อภาคเรียน"],
    ["min_hours_per_year", "50", "เกณฑ์ชั่วโมงขั้นต่ำต่อปีการศึกษา"],
    ["max_hours_scale", "400", "เพดานชั่วโมงสะสมสูงสุด"],
    ["academic_year", "2569", "ปีการศึกษา"],
    ["security_mode", "OWNER_ONLY_SERVER_SECRETS", "ห้ามเก็บ credential ในไฟล์หรือ frontend"],
]

FORBIDDEN_KEYS = {
    "password", "admin_password", "teacher_password", "token", "bot_token",
    "telegram_bot_token", "telegram_chat_id", "api_key", "secret", "line_user_id",
}


def add_or_update_settings_sheet(file_path):
    if not os.path.exists(file_path):
        return
    wb = openpyxl.load_workbook(file_path)
    sheet_name = "ตั้งค่าระบบ" if "ตั้งค่าระบบ" in wb.sheetnames else "Settings"
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)
    ws.delete_rows(1, ws.max_row + 1)
    for row in SETTINGS_DATA:
        key = str(row[0]).strip().lower()
        if key in FORBIDDEN_KEYS or any(word in key for word in ("password", "token", "secret", "api_key")):
            raise RuntimeError(f"Refusing to write secret setting key: {key}")
        ws.append(row)
    wb.save(file_path)
    wb.close()


def main():
    for path in EXCEL_FILES:
        add_or_update_settings_sheet(path)


if __name__ == "__main__":
    main()
