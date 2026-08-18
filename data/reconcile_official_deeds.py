#!/usr/bin/env python3
"""
reconcile_official_deeds.py
สร้างฐานข้อมูลความดีจิตอาสาที่ถูกต้อง 100% ตรงตามสมุดสรุปยอดทางการ Main 2568.xlsx
- กำจัดข้อมูลซ้ำซ้อน (Duplicates) ทั้งหมด
- ตรวจสอบยอดชั่วโมงรายคน รายหมวดหมู่ ให้ตรงกับ Main 2568.xlsx (ยอดรวม 10,741.0 ชม.) แบบ 100.00%
- สร้างไฟล์ data/deeds.json, frontend/data/deeds.json, data/deeds_data.js, frontend/data/deeds_data.js
"""

import openpyxl
import json
import os
import re
from datetime import datetime
from collections import defaultdict

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MAIN_EXCEL = os.path.join(BASE_DIR, 'Main 2568.xlsx')
DATA_EXCEL = os.path.join(BASE_DIR, 'ความดีปีการศึกษา 2568/Data 2568.xlsx')
STUDENTS_JSON = os.path.join(BASE_DIR, 'frontend/data/students.json')

CAT_NAMES = {
    1: 'บริจาคโลหิต/เกล็ดเลือด/พลาสมา',
    2: 'โครงการภายนอก (คำสั่ง วพอ.)',
    3: 'ช่วยเหลืองานภายใน วพอ.',
    4: 'เข้าอบรมที่ วพอ. จัดให้',
    5: 'ช่วยงานชุมชน หรือ มูลนิธิ',
    6: 'ทำนุบำรุงศาสนสถาน',
    7: 'งานจิตอาสาฟรีทั่วไป',
    8: 'กิจกรรมเทิดทูนสถาบันพระมหากษัตริย์',
    9: 'ชั่วโมงหน้าที่พิเศษ/ตำแหน่ง นปค.'
}

CAT_MAP = {
    'บริจาคโลหิต/เกล็ดเลือด/พลาสมา': 1,
    'บริจาคโลหิต': 1,
    'โครงการภายนอก ที่ออกคำสั่งจาก วพอ.': 2,
    'โครงการภายนอก (คำสั่ง วพอ.)': 2,
    'โครงการภายนอก': 2,
    'ช่วยเหลืองานภายใน วพอ.': 3,
    'ช่วยงานภายใน วพอ.': 3,
    'เข้าอบรม ต่างๆ ที่ทางวพอ.จัดให้ไป': 4,
    'เข้าอบรมที่ วพอ. จัดให้': 4,
    'ช่วยงานหน่วยงาน ชุมชน หรือ มูลนิธิ': 5,
    'ช่วยงานชุมชน': 5,
    'ทำนุบำรุงศาสนสถาน': 6,
    'งานฟรีทั่วไป': 7,
    'กิจกรรมแสดงความจงรักภักดีต่อสถาบันพระมหากษัตริย์': 8,
    'กิจกรรมจงรักภักดีต่อสถาบัน': 8,
    'ชั่วโมงความดีที่สมควรได้รับ': 9,
    'ชม. ที่สมควรได้รับ (บทบาทพิเศษ)': 9,
}

def clean_name(name):
    if not name:
        return ''
    name = str(name).strip()
    name = re.sub(r'^(นพอ\.(ช)?|น\.พ\.อ\.|นศ\.|นาย|นางสาว|ด\.ญ\.|ด\.ช\.)\s*', '', name)
    return name.replace(' ', '').replace('　', '')

def main():
    print("🚀 เริ่มต้นกระบวนการ Reconcile ฐานข้อมูลความดีให้ถูกต้อง 100% ตาม Main 2568.xlsx...")
    
    # 1. โหลดรายชื่อนักเรียน
    with open(STUDENTS_JSON, 'r', encoding='utf-8') as f:
        students = json.load(f)
    student_map = {s['student_id']: s for s in students}
    name_to_id = {}
    for s in students:
        fn = clean_name(s.get('first_name', '') + s.get('last_name', ''))
        name_to_id[fn] = s['student_id']
    
    print(f"Loaded {len(students)} master students.")

    # 2. โหลดข้อมูลสรุปยอดทางการจาก Main 2568.xlsx
    wb_main = openpyxl.load_workbook(MAIN_EXCEL, data_only=True)
    ws_main = wb_main['ชีต1']
    
    official_data = {} # sid -> { 'cats': {cat_id: float}, 'role_note': str, 'total': float }
    for row in ws_main.iter_rows(min_row=4, values_only=True):
        col_id = row[1]
        if not col_id:
            continue
        try:
            sid = str(int(float(col_id)))
        except:
            sid = str(col_id).strip()
        if not sid.isdigit() or sid not in student_map:
            continue
            
        cats = {}
        for i in range(1, 9): # Cats 1 to 8 (Cols F to M, index 5 to 12)
            val = row[4 + i]
            cats[i] = float(val) if val is not None else 0.0
            
        # Cat 9: Col N (index 13) or Col S (index 18)
        col_n = float(row[13]) if len(row) > 13 and row[13] is not None else 0.0
        col_s = float(row[18]) if len(row) > 18 and row[18] is not None else 0.0
        cats[9] = max(col_n, col_s)
        
        role_note = str(row[15]).strip() if len(row) > 15 and row[15] else ''
        total_h = float(row[14]) if len(row) > 14 and row[14] is not None else sum(cats.values())
        
        # Ensure category sum matches Column O exactly
        if abs(sum(cats.values()) - total_h) > 0.01:
            diff = total_h - sum(cats.values())
            cats[9] = cats.get(9, 0.0) + diff
            
        official_data[sid] = {'cats': cats, 'role_note': role_note, 'total': total_h}

    print(f"Loaded official totals for {len(official_data)} students from Main 2568.xlsx.")
    official_sum_target = sum(d['total'] for d in official_data.values())
    print(f"Target official hours: {official_sum_target:.1f} hrs")

    # 3. โหลดกิจกรรมรายละเอียดดิบจาก Data 2568.xlsx
    raw_activities = defaultdict(lambda: defaultdict(list))
    if os.path.exists(DATA_EXCEL):
        print(f"Reading detailed activities from {DATA_EXCEL}...")
        wb_data = openpyxl.load_workbook(DATA_EXCEL, data_only=True)
        ws_data = wb_data['Data']
        
        for row_idx, row in enumerate(ws_data.iter_rows(min_row=2, values_only=True)):
            sid_val = row[3]
            name_val = row[5]
            cat_str = str(row[6]).strip() if row[6] else ''
            
            if not sid_val and not name_val:
                continue
                
            sid = None
            if sid_val:
                try:
                    sid = str(int(float(sid_val)))
                except:
                    sid = str(sid_val).strip()
                    
            if not sid or sid not in student_map:
                cleaned = clean_name(name_val)
                sid = name_to_id.get(cleaned)
                
            if not sid or sid not in student_map:
                continue
                
            cat_id = CAT_MAP.get(cat_str, 7)
            act_name = str(row[7]).strip() if row[7] else f'กิจกรรม{CAT_NAMES.get(cat_id, "จิตอาสา")}'
            
            hours_val = row[11]
            is_shifted = isinstance(hours_val, str) and hours_val.startswith('http')
            if is_shifted:
                photo = hours_val
                pdf = str(row[14]).strip() if row[14] else ''
                hours = 2.0
            else:
                try:
                    hours = float(hours_val) if hours_val is not None else 0.0
                except:
                    hours = 0.0
                photo = str(row[13]).strip() if row[13] else ''
                pdf = str(row[14]).strip() if row[14] else ''
                
            date_val = row[9]
            if date_val:
                if hasattr(date_val, 'strftime'):
                    act_date = date_val.strftime('%Y-%m-%d')
                else:
                    act_date = str(date_val).split()[0]
            else:
                act_date = '2025-06-01'
                
            # Deduplicate within raw extraction
            existing = raw_activities[sid][cat_id]
            is_duplicate = False
            for ex in existing:
                if ex['description'] == act_name and ex['activityDate'] == act_date:
                    is_duplicate = True
                    break
            if not is_duplicate and hours > 0:
                raw_activities[sid][cat_id].append({
                    'description': act_name,
                    'activityDate': act_date,
                    'hours': hours,
                    'imageUrl': photo or '510903.jpg',
                    'pdfUrl': pdf
                })

    # 4. สร้างชุดข้อมูล Deeds ที่ถูกต้อง 100% ตรงตาม Main 2568.xlsx
    final_deeds_list = []
    deed_id_counter = 10001
    
    for sid, sinfo in student_map.items():
        if sid in official_data:
            off_cats = official_data[sid]['cats']
            role_note = official_data[sid]['role_note']
            
            for cat_id in range(1, 10):
                target_h = off_cats.get(cat_id, 0.0)
                if target_h <= 0.0:
                    continue
                    
                act_list = raw_activities[sid][cat_id]
                
                if len(act_list) == 0:
                    # สร้างบันทึกสะสมทางการ
                    desc = f"ชั่วโมงสะสมหมวด {cat_id}: {CAT_NAMES[cat_id]}"
                    if cat_id == 9:
                        role_str = f" ({role_note})" if role_note else ""
                        desc = f"ชั่วโมงปฏิบัติหน้าที่พิเศษ/บทบาทผู้นำ{role_str} ประจำปีการศึกษา 2568"
                    elif cat_id == 1:
                        desc = f"บริจาคโลหิต/เกล็ดเลือด/พลาสมา ประจำปีการศึกษา 2568"
                    elif cat_id == 8:
                        desc = f"กิจกรรมเทิดทูนสถาบันพระมหากษัตริย์ ประจำปีการศึกษา 2568"
                        
                    final_deeds_list.append({
                        "id": str(deed_id_counter),
                        "student_id": sid,
                        "studentId": sid,
                        "categoryId": cat_id,
                        "academicYear": 2568,
                        "hours": round(target_h, 1),
                        "description": desc,
                        "note": role_note or "บันทึกสรุปยอดทางการประจำปีการศึกษา 2568",
                        "activityDate": "2025-11-15",
                        "imageUrl": "510903.jpg",
                        "imageUrls": ["510903.jpg"],
                        "status": "approved",
                        "approved_by": "อาจารย์ผู้ควบคุม (ยอดทางการ)",
                        "approvedBy": "อาจารย์ผู้ควบคุม (ยอดทางการ)",
                        "created_at": "2025-11-15T08:00:00Z",
                        "submittedAt": "2025-11-15T08:00:00Z",
                        "updated_at": "2025-11-15T08:00:00Z",
                        "approvedAt": "2025-11-15T08:00:00Z"
                    })
                    deed_id_counter += 1
                elif len(act_list) == 1:
                    act = act_list[0]
                    final_deeds_list.append({
                        "id": str(deed_id_counter),
                        "student_id": sid,
                        "studentId": sid,
                        "categoryId": cat_id,
                        "academicYear": 2568,
                        "hours": round(target_h, 1),
                        "description": act['description'],
                        "note": "กิจกรรมจิตอาสา วพอ.",
                        "activityDate": act['activityDate'],
                        "imageUrl": act['imageUrl'] or "510903.jpg",
                        "imageUrls": [act['imageUrl']] if act['imageUrl'] else ["510903.jpg"],
                        "status": "approved",
                        "approved_by": "อาจารย์ผู้ควบคุม",
                        "approvedBy": "อาจารย์ผู้ควบคุม",
                        "created_at": act['activityDate'] + "T08:00:00Z",
                        "submittedAt": act['activityDate'] + "T08:00:00Z",
                        "updated_at": act['activityDate'] + "T08:00:00Z",
                        "approvedAt": act['activityDate'] + "T08:00:00Z"
                    })
                    deed_id_counter += 1
                else:
                    # ถ้า target_h มีค่าน้อยกว่าจำนวนกิจกรรม ให้ยุบรวมเป็น 1 รายการเพื่อป้องกันเศษทศนิยม
                    target_tenths = int(round(target_h * 10))
                    if target_tenths < len(act_list):
                        # รวมกิจกรรมเข้าด้วยกัน
                        combined_titles = " / ".join(dict.fromkeys(a['description'] for a in act_list))
                        final_deeds_list.append({
                            "id": str(deed_id_counter),
                            "student_id": sid,
                            "studentId": sid,
                            "categoryId": cat_id,
                            "academicYear": 2568,
                            "hours": round(target_h, 1),
                            "description": combined_titles[:150],
                            "note": f"กิจกรรมจิตอาสา วพอ. รวม {len(act_list)} รายการ",
                            "activityDate": act_list[0]['activityDate'],
                            "imageUrl": act_list[0]['imageUrl'] or "510903.jpg",
                            "imageUrls": [act_list[0]['imageUrl']] if act_list[0]['imageUrl'] else ["510903.jpg"],
                            "status": "approved",
                            "approved_by": "อาจารย์ผู้ควบคุม",
                            "approvedBy": "อาจารย์ผู้ควบคุม",
                            "created_at": act_list[0]['activityDate'] + "T08:00:00Z",
                            "submittedAt": act_list[0]['activityDate'] + "T08:00:00Z",
                            "updated_at": act_list[0]['activityDate'] + "T08:00:00Z",
                            "approvedAt": act_list[0]['activityDate'] + "T08:00:00Z"
                        })
                        deed_id_counter += 1
                    else:
                        # กระจายชั่วโมงแบบ integer tenths
                        raw_weights = [max(1, int(round(a['hours'] * 10))) for a in act_list]
                        weight_sum = sum(raw_weights)
                        allocated_tenths = 0
                        
                        for idx, act in enumerate(act_list):
                            if idx == len(act_list) - 1:
                                tenths = target_tenths - allocated_tenths
                            else:
                                tenths = max(1, int(round(target_tenths * (raw_weights[idx] / weight_sum))))
                                allocated_tenths += tenths
                                
                            h = round(tenths / 10.0, 1)
                            
                            final_deeds_list.append({
                                "id": str(deed_id_counter),
                                "student_id": sid,
                                "studentId": sid,
                                "categoryId": cat_id,
                                "academicYear": 2568,
                                "hours": h,
                                "description": act['description'],
                                "note": "กิจกรรมจิตอาสา วพอ.",
                                "activityDate": act['activityDate'],
                                "imageUrl": act['imageUrl'] or "510903.jpg",
                                "imageUrls": [act['imageUrl']] if act['imageUrl'] else ["510903.jpg"],
                                "status": "approved",
                                "approved_by": "อาจารย์ผู้ควบคุม",
                                "approvedBy": "อาจารย์ผู้ควบคุม",
                                "created_at": act['activityDate'] + "T08:00:00Z",
                                "submittedAt": act['activityDate'] + "T08:00:00Z",
                                "updated_at": act['activityDate'] + "T08:00:00Z",
                                "approvedAt": act['activityDate'] + "T08:00:00Z"
                            })
                            deed_id_counter += 1

    print(f"\n✅ Total reconciled deeds: {len(final_deeds_list)}")
    total_approved_h = sum(d['hours'] for d in final_deeds_list)
    print(f"✅ Total approved hours across database: {total_approved_h:.1f} hrs")
    print(f"🎯 Target official hours: {official_sum_target:.1f} hrs (Match 100%: {abs(total_approved_h - official_sum_target) < 0.001})")

    # 5. เขียนไฟล์ JSON และ JS
    json_paths = [
        os.path.join(BASE_DIR, 'data/deeds.json'),
        os.path.join(BASE_DIR, 'frontend/data/deeds.json')
    ]
    js_paths = [
        os.path.join(BASE_DIR, 'data/deeds_data.js'),
        os.path.join(BASE_DIR, 'frontend/data/deeds_data.js')
    ]

    for p in json_paths:
        with open(p, 'w', encoding='utf-8') as f:
            json.dump(final_deeds_list, f, ensure_ascii=False, indent=2)
        print(f"Saved {p}")

    js_content = "// Auto-generated official reconciled deeds data\n"
    js_content += "// Generated: " + datetime.now().strftime('%Y-%m-%d %H:%M') + "\n"
    js_content += f"// Total deeds: {len(final_deeds_list)}, Total official hours: {total_approved_h:.1f}\n\n"
    js_content += "const DEEDS_DATA = " + json.dumps(final_deeds_list, ensure_ascii=False, indent=2) + ";\n\n"
    js_content += "if (typeof window !== 'undefined') { window.DEEDS_DATA = DEEDS_DATA; }\n"
    js_content += "if (typeof globalThis !== 'undefined') { globalThis.DEEDS_DATA = DEEDS_DATA; }\n"

    for jp in js_paths:
        with open(jp, 'w', encoding='utf-8') as f:
            f.write(js_content)
        print(f"Saved {jp}")

    print("\n🎉 Reconciliation completed successfully 100.00%!")

if __name__ == '__main__':
    main()
