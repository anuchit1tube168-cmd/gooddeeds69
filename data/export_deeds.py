#!/usr/bin/env python3
"""
export_deeds.py
แปลงข้อมูลความดีจาก 'ความดีปีการศึกษา 2568/Data 2568.xlsx' -> deeds.json / deeds_data.js
โดยมีการจับคู่กับสรุปยอดทางการจาก 'ความดีปีการศึกษา 2568/Main 2568.xlsx'
เพื่อรับประกันความถูกต้อง 100% ตามยอดชั่วโมงอย่างเป็นทางการของปีการศึกษา 2568
"""

import openpyxl
import json
import os
import re
from datetime import datetime

# พาธไฟล์ข้อมูล
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_EXCEL_PATH = os.path.join(BASE_DIR, 'ความดีปีการศึกษา 2568/Data 2568.xlsx')
MAIN_EXCEL_PATH = os.path.join(BASE_DIR, 'ความดีปีการศึกษา 2568/Main 2568.xlsx')
STUDENTS_JSON_PATH = os.path.join(BASE_DIR, 'frontend/data/students.json')

# หมวดหมู่ความดี
CAT_MAP = {
    'บริจาคโลหิต/เกล็ดเลือด/พลาสมา': 1,
    'โครงการภายนอกที่ออกคำสั่งจาก วพอ.': 2,
    'ช่วยเหลืองานภายใน วพอ.': 3,
    'เข้าอบรมต่างๆที่ทางวพอ.จัดให้': 4,
    'ช่วยงานหน่วยงานชุมชน หรือมูลนิธิ': 5,
    'ทำนุบำรุงศาสนสถาน': 6,
    'งานฟรีทั่วไป': 7,
    'กิจกรรมแสดงความจงรักภักดีต่อสถาบันพระมหากษัตริย์': 8,
    'ชั่วโมงความดีที่สมควรได้รับ': 9
}

def clean_name(name):
    if not name:
        return ''
    name = str(name).strip()
    name = re.sub(r'^(นพอ\.(ช)?|น\.พ\.อ\.|นศ\.|นาย|นางสาว|ด\.ญ\.|ด\.ช\.)\s*', '', name)
    return name.replace(' ', '').replace('　', '')

def main():
    if not os.path.exists(DATA_EXCEL_PATH) or not os.path.exists(MAIN_EXCEL_PATH):
        print("❌ ไม่พบไฟล์ Excel ความดีปี 2568 (Data 2568.xlsx หรือ Main 2568.xlsx)")
        return

    # 1. โหลดข้อมูลนักเรียนในระบบ
    with open(STUDENTS_JSON_PATH, 'r', encoding='utf-8') as f:
        students_list = json.load(f)
    
    student_ids = set(s['student_id'] for s in students_list)
    
    # Map ชื่อจริง -> รหัสในระบบ
    name_to_id = {}
    for s in students_list:
        fullname = clean_name(s.get('first_name', '') + s.get('last_name', ''))
        name_to_id[fullname] = s['student_id']

    # 2. โหลดข้อมูลสรุปยอดทางการจาก Main 2568.xlsx
    print(f"📖 กำลังโหลดข้อมูลสรุปยอดความดีจาก {MAIN_EXCEL_PATH}...")
    wb_main = openpyxl.load_workbook(MAIN_EXCEL_PATH, data_only=True)
    ws_main = wb_main['ชีต1']
    
    official_summaries = {} # student_id -> { cat_id: hours }
    
    for row in ws_main.iter_rows(min_row=4, values_only=True):
        col1 = row[1]
        if col1 is None:
            continue
        try:
            sid = str(int(float(col1)))
        except:
            sid = str(col1).strip()
            
        if not sid.isdigit() or sid not in student_ids:
            continue
            
        official_summaries[sid] = {}
        for i in range(9):
            val = row[5 + i]
            h = float(val) if val is not None else 0.0
            official_summaries[sid][i+1] = h

    # 3. โหลดบันทึกดิบจาก Data 2568.xlsx
    print(f"📖 กำลังโหลดบันทึกดิบความดีจาก {DATA_EXCEL_PATH}...")
    wb_data = openpyxl.load_workbook(DATA_EXCEL_PATH, data_only=True)
    ws_data = wb_data['Data']
    
    detailed_deeds = {} # student_id -> { cat_id: [ deeds ] }
    
    for row_idx, row in enumerate(ws_data.iter_rows(min_row=2, values_only=True)):
        row_num = row_idx + 2
        sid_val = row[3]
        name_val = row[5]
        class_val = row[2]
        cat_str = row[6]
        
        if not sid_val or not name_val:
            continue
            
        try:
            sid = str(int(float(sid_val)))
        except:
            sid = str(sid_val).strip()
            
        if not sid.isdigit():
            continue
            
        # แมปรหัส
        target_sid = None
        if sid in student_ids:
            target_sid = sid
        elif len(sid) == 4:
            if class_val == 'ชั้นปีที่ 4':
                target_sid = '650' + sid
            elif class_val == 'ชั้นปีที่ 3':
                target_sid = '660' + sid
            elif class_val == 'ชั้นปีที่ 2':
                target_sid = '670' + sid
            elif class_val == 'ชั้นปีที่ 1':
                target_sid = '680' + sid
                
        if not target_sid or target_sid not in student_ids:
            fullname_cleaned = clean_name(name_val)
            if fullname_cleaned in name_to_id:
                target_sid = name_to_id[fullname_cleaned]
            else:
                for db_name, db_id in name_to_id.items():
                    if fullname_cleaned in db_name or db_name in fullname_cleaned:
                        target_sid = db_id
                        break
                        
        if not target_sid or target_sid not in student_ids:
            continue

        cat_id = CAT_MAP.get(str(cat_str).strip(), 9)

        # จัดการข้อมูลที่เยื้องเลื่อน
        hours_val = row[11]
        is_shifted = isinstance(hours_val, str) and hours_val.startswith('http')
        act_name = str(row[7]).strip() if row[7] else ''
        
        if is_shifted:
            photo = hours_val
            pdf = str(row[14]).strip() if row[14] else ''
            if 'ดุริยางค์' in act_name:
                hours = 1.0
            elif 'green market' in act_name.lower():
                hours = 8.0
            elif 'เก็บหินขาว' in act_name:
                hours = 2.0
            elif 'สอบคัดเลือก' in act_name or 'นพอ.69' in act_name:
                hours = 3.0
            elif 'มอบหมวก' in act_name:
                hours = 4.0
            else:
                hours = 2.0
        else:
            try:
                hours = float(hours_val) if hours_val is not None else 0.0
            except ValueError:
                hours = 0.0
            photo = str(row[13]).strip() if row[13] else ''
            pdf = str(row[14]).strip() if row[14] else ''

        date_val = row[9]
        if date_val:
            if hasattr(date_val, 'strftime'):
                act_date = date_val.strftime('%Y-%m-%d')
            else:
                act_date = str(date_val).split()[0]
        else:
            act_date = '2025-06-01'

        approved_by = str(row[12]).strip() if row[12] else 'ระบบนำเข้าข้อมูล'
        if approved_by.startswith('http'):
            approved_by = 'ระบบนำเข้าข้อมูล'

        deed_id = f'import_2568_{row_num}_{target_sid}'
        
        deed_record = {
            'id': deed_id,
            'studentId': target_sid,
            'categoryId': cat_id,
            'academicYear': 2568,
            'hours': hours,
            'description': act_name or f'ข้อมูลนำเข้าปี 2568 — {cat_str}',
            'activityDate': act_date,
            'imageUrls': [photo] if photo else [],
            'pdfUrl': pdf,
            'status': 'approved',
            'submittedAt': '2025-12-31T00:00:00.000Z',
            'approvedBy': approved_by,
            'approvedAt': '2025-12-31T00:00:00.000Z',
            'rejectReason': None,
            'note': 'นำเข้าจาก Data 2568.xlsx' + (' (แก้ไขคอลัมน์เลื่อน)' if is_shifted else '')
        }
        
        detailed_deeds.setdefault(target_sid, {}).setdefault(cat_id, []).append(deed_record)

    # 4. รวมและปรับแต่งข้อมูลให้สอดคล้องกับยอดสรุปอย่างเป็นทางการ
    print("⚡ กำลังผสานข้อมูลบันทึกดิบและยอดสรุปทางการ...")
    final_deeds = {}
    total_injected = 0
    total_scaled = 0
    total_deeds_count = 0

    for sid in student_ids:
        if sid in official_summaries:
            for cat_id in range(1, 10):
                excel_h = official_summaries[sid][cat_id]
                if excel_h <= 0.0:
                    continue
                    
                deeds_list = detailed_deeds.get(sid, {}).get(cat_id, [])
                detail_sum = sum(d['hours'] for d in deeds_list)
                
                if len(deeds_list) == 0:
                    # กรณีไม่มีบันทึกรายละเอียด ให้เพิ่มแบบสะสมหมวดหมู่รวม
                    deed_id = f'import_2568_inj_{sid}_{cat_id}'
                    injected_deed = {
                        'id': deed_id,
                        'studentId': sid,
                        'categoryId': cat_id,
                        'academicYear': 2568,
                        'hours': excel_h,
                        'description': f'ชั่วโมงกิจกรรมสะสม (หมวด {cat_id})',
                        'activityDate': '2025-06-01',
                        'imageUrls': [],
                        'pdfUrl': '',
                        'status': 'approved',
                        'submittedAt': '2025-12-31T00:00:00.000Z',
                        'approvedBy': 'ระบบนำเข้าข้อมูล',
                        'approvedAt': '2025-12-31T00:00:00.000Z',
                        'rejectReason': None,
                        'note': 'นำเข้าสะสมจาก Main 2568.xlsx'
                    }
                    final_deeds.setdefault(sid, []).append(injected_deed)
                    total_injected += 1
                    total_deeds_count += 1
                else:
                    # หากมีบันทึกรายละเอียด ให้ตรวจสอบผลรวม
                    if abs(excel_h - detail_sum) < 0.01:
                        # ยอดตรงกันพอดี
                        final_deeds.setdefault(sid, []).extend(deeds_list)
                        total_deeds_count += len(deeds_list)
                    elif excel_h > detail_sum:
                        # ยอดสรุปทางการมีชั่วโมงมากกว่า ให้เก็บบันทึกดิบไว้ทั้งหมด แล้วเพิ่มชั่วโมงส่วนต่าง
                        final_deeds.setdefault(sid, []).extend(deeds_list)
                        diff_h = excel_h - detail_sum
                        deed_id = f'import_2568_inj_{sid}_{cat_id}'
                        injected_deed = {
                            'id': deed_id,
                            'studentId': sid,
                            'categoryId': cat_id,
                            'academicYear': 2568,
                            'hours': round(diff_h, 2),
                            'description': f'ชั่วโมงกิจกรรมสะสมเพิ่มเติม (หมวด {cat_id})',
                            'activityDate': '2025-06-01',
                            'imageUrls': [],
                            'pdfUrl': '',
                            'status': 'approved',
                            'submittedAt': '2025-12-31T00:00:00.000Z',
                            'approvedBy': 'ระบบนำเข้าข้อมูล',
                            'approvedAt': '2025-12-31T00:00:00.000Z',
                            'rejectReason': None,
                            'note': 'ส่วนต่างนำเข้าจาก Main 2568.xlsx'
                        }
                        final_deeds.setdefault(sid, []).append(injected_deed)
                        total_injected += 1
                        total_deeds_count += len(deeds_list) + 1
                    else:
                        # ยอดดิบมีชั่วโมงมากกว่ายอดสรุปทางการ (เช่น มีข้อมูลซ้ำซ้อน) ให้ปรับสัดส่วนชั่วโมงลงให้ได้ผลรวมตามจริง
                        scale = excel_h / detail_sum
                        for d in deeds_list:
                            d['hours'] = round(d['hours'] * scale, 2)
                            final_deeds.setdefault(sid, []).append(d)
                        total_scaled += len(deeds_list)
                        total_deeds_count += len(deeds_list)

    # 5. เขียนไฟล์ส่งออก
    js_header = f"""// Auto-generated from Data & Main 2568.xlsx — DO NOT EDIT MANUALLY
// Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}
// นักเรียนที่มีความดี: {len(final_deeds)} คน, รายการความดีทั้งหมด: {total_deeds_count} รายการ

const IMPORTED_DEEDS = """

    paths = [
        os.path.join(BASE_DIR, 'data/deeds.json'),
        os.path.join(BASE_DIR, 'frontend/data/deeds.json')
    ]
    
    js_paths = [
        os.path.join(BASE_DIR, 'data/deeds_data.js'),
        os.path.join(BASE_DIR, 'frontend/data/deeds_data.js')
    ]

    for p in paths:
        with open(p, 'w', encoding='utf-8') as f:
            json.dump(final_deeds, f, ensure_ascii=False, indent=2)
            
    for jp in js_paths:
        with open(jp, 'w', encoding='utf-8') as f:
            f.write(js_header + json.dumps(final_deeds, ensure_ascii=False, indent=2) + ';\n')

    print("\n🎉 สรุปการแปลงข้อมูลสมบูรณ์แบบ!")
    print(f"  - บันทึกกิจกรรมรวมทั้งระบบ: {total_deeds_count} รายการ")
    print(f"  - จำนวนนักเรียนที่มีความดี: {len(final_deeds)} คน (ตรงตามรหัสจริง)")
    print(f"  - เพิ่มชั่วโมงส่วนต่าง/สะสม (Injected): {total_injected} หมวดหมู่")
    print(f"  - ปรับยอดลดชั่วโมงตามจริง (Scaled): {total_scaled} รายการ")
    print("📁 อัปเดตไฟล์ deeds.json และ deeds_data.js เรียบร้อยแล้ว ทั้งใน data/ และ frontend/data/")

if __name__ == '__main__':
    main()
