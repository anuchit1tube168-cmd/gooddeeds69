#!/usr/bin/env python3
"""
inspect_2568_deeds.py
ตรวจสอบโครงสร้างไฟล์ Excel บันทึกความดีปีการศึกษา 2568 ทั้ง 5 ไฟล์
"""
import openpyxl
import os

FILES = [
    "/Users/agislamious/Downloads/Main 2568.xlsx",
    "/Users/agislamious/Downloads/Data 2568.xlsx",
    "/Users/agislamious/Downloads/ความดีปีการศึกษา 2568/สรุปยอดบันทึกความดีนพอ.ชั้นปีที่3.xlsx",
    "/Users/agislamious/Downloads/ความดีปีการศึกษา 2568/รวมบันทึกความดีปี 68 ชั้นปีที่1.xlsx",
    "/Users/agislamious/Downloads/ความดีปีการศึกษา 2568/รวมบันทึกความดีปีการศึกษา 68 ชั้นปีที่ 2.xlsx",
]

def inspect_file(fpath):
    if not os.path.exists(fpath):
        print(f"❌ ไม่พบไฟล์: {fpath}")
        return
    print(f"\n==================================================")
    print(f"📄 ไฟล์: {os.path.basename(fpath)}")
    print(f"==================================================")
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        print(f"Sheet names: {wb.sheetnames}")
        for sname in wb.sheetnames:
            ws = wb[sname]
            print(f"\n  📊 Sheet '{sname}': {ws.max_row} rows x {ws.max_column} cols")
            for r_idx in range(1, min(10, ws.max_row + 1)):
                row_vals = [ws.cell(r_idx, c_idx).value for c_idx in range(1, min(10, ws.max_column + 1))]
                if any(row_vals):
                    print(f"    R{r_idx}: {row_vals}")
        wb.close()
    except Exception as e:
        print(f"⚠️ Error reading {fpath}: {e}")

def main():
    for f in FILES:
        inspect_file(f)

if __name__ == '__main__':
    main()
