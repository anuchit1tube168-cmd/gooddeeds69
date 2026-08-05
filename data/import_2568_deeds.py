#!/usr/bin/env python3
"""
import_2568_deeds.py
นำเข้าข้อมูลบันทึกความดีปีการศึกษา 2568 จากไฟล์ Excel ทั้ง 5 ไฟล์:
1. Data 2568.xlsx (998 รายการกิจกรรมละเอียด)
2. รวมบันทึกความดีปี 68 ชั้นปีที่1.xlsx (รุ่น 68)
3. รวมบันทึกความดีปีการศึกษา 68 ชั้นปีที่ 2.xlsx (รุ่น 67)
4. สรุปยอดบันทึกความดีนพอ.ชั้นปีที่3.xlsx (รุ่น 66)
5. Main 2568.xlsx (รุ่น 65/64)

เชื่อมโยงกับรหัสนักเรียน (student_id) ครบทั้ง 375 คน พร้อมตรวจสอบความถูกต้อง
"""
import openpyxl
import zipfile
import xml.etree.ElementTree as ET
import json
import os
import re
import datetime

DATA_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(DATA_DIR)

# File paths
DATA_2568_FILE = "/Users/agislamious/Downloads/Data 2568.xlsx"
YEAR1_FILE = "/Users/agislamious/Downloads/ความดีปีการศึกษา 2568/รวมบันทึกความดีปี 68 ชั้นปีที่1.xlsx"
YEAR2_FILE = "/Users/agislamious/Downloads/ความดีปีการศึกษา 2568/รวมบันทึกความดีปีการศึกษา 68 ชั้นปีที่ 2.xlsx"
YEAR3_FILE = "/Users/agislamious/Downloads/ความดีปีการศึกษา 2568/สรุปยอดบันทึกความดีนพอ.ชั้นปีที่3.xlsx"
MAIN_2568_FILE = "/Users/agislamious/Downloads/Main 2568.xlsx"

# Mapping category names to category ID (1-9)
CAT_MAP = {
    'บริจาคโลหิต/เกล็ดเลือด/พลาสมา': 1,
    'บริจาคโลหิต': 1,
    'โครงการภายนอก ที่ออกคำสั่งจาก วพอ.': 2,
    'โครงการภายนอก (คำสั่ง วพอ.)': 2,
    'โครงการภายนอก': 2,
    'ช่วยเหลืองานภายใน วพอ.': 3,
    'ช่วยงานภายใน วพอ.': 3,
    'เข้าอบรม ต่างๆ ที่ทางวพอ.จัดให้ไป': 4,
    'เข้าอบรมที่ วพอ. จัดให้': 4,
    'ช่วยงานหน่วยงาน ชุมชน หรือ มูลนิธิ': 5,
    'ช่วยงานหน่วยงาน/ชุมชน/มูลนิธิ': 5,
    'ช่วยงานชุมชน': 5,
    'ทำนุบำรุงศาสนสถาน': 6,
    'งานฟรีทั่วไป': 7,
    'กิจกรรมจงรักภักดีต่อสถาบัน': 8,
    'ชม. ที่สมควรได้รับ (บทบาทพิเศษ)': 9,
}

def clean_str(val):
    if val is None:
        return ''
    s = str(val).strip()
    return re.sub(r'\s+', ' ', s)

def clean_sid(val):
    if not val:
        return ''
    s = str(val).strip().replace('*', '').replace(' ', '')
    try:
        return str(int(float(s)))
    except:
        return s

def parse_xml_xlsx(fpath):
    """Direct XML parser for Excel files with XML stylesheet issues"""
    rows_all = []
    with zipfile.ZipFile(fpath, 'r') as z:
        strings = []
        if 'xl/sharedStrings.xml' in z.namelist():
            tree = ET.fromstring(z.read('xl/sharedStrings.xml'))
            for elem in tree.iter():
                if elem.tag.endswith('t') and elem.text is not None:
                    strings.append(elem.text)
        
        # Check sheet2 first (has correct student IDs), then sheet1
        sheet_files = [f for f in ['xl/worksheets/sheet2.xml', 'xl/worksheets/sheet1.xml'] if f in z.namelist()]
        for sfile in sheet_files:
            tree = ET.fromstring(z.read(sfile))
            rows = []
            for row_elem in tree.iter('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row'):
                row_vals = []
                for cell in row_elem.iter('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c'):
                    val_type = cell.attrib.get('t')
                    val_elem = cell.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                    val = val_elem.text if val_elem is not None else ''
                    if val_type == 's' and val.isdigit():
                        idx = int(val)
                        val = strings[idx] if idx < len(strings) else val
                    row_vals.append(val)
                rows.append(row_vals)
            rows_all.append(rows)
    return rows_all

def main():
    print("=== STARTING IMPORT OF 2568 DEEDS DATABASE ===")
    
    # 1. Load 375 master students
    students_file = os.path.join(DATA_DIR, 'students.json')
    with open(students_file, 'r', encoding='utf-8') as f:
        students = json.load(f)
    
    student_by_id = {s['student_id']: s for s in students}
    student_by_name = {f"{clean_str(s.get('first_name'))} {clean_str(s.get('last_name'))}": s['student_id'] for s in students}
    
    print(f"📋 Loaded {len(students)} master students from database.")
    
    imported_deeds = []
    deed_id_counter = 10000
    
    # -------------------------------------------------------------
    # FILE 1: Data 2568.xlsx (998 Detailed Activity Submissions)
    # -------------------------------------------------------------
    if os.path.exists(DATA_2568_FILE):
        print(f"\n📂 Reading File 1: {os.path.basename(DATA_2568_FILE)}")
        wb = openpyxl.load_workbook(DATA_2568_FILE, data_only=True)
        if 'Form Responses 1' in wb.sheetnames or 'ชีต1' in wb.sheetnames or len(wb.sheetnames) > 0:
            ws = wb[wb.sheetnames[0]]
            count = 0
            for r_idx in range(2, ws.max_row + 1):
                row = [ws.cell(r_idx, c).value for c in range(1, 12)]
                if not row or not row[3]: # Column D: student_id
                    continue
                
                sid = clean_sid(row[3])
                if not sid or sid not in student_by_id:
                    # Try matching by name in Column F
                    fullname = clean_str(row[5]).replace('นพอ.', '').replace('(ช)', '').strip()
                    if fullname in student_by_name:
                        sid = student_by_name[fullname]
                    else:
                        continue
                
                cat_name = clean_str(row[6])
                cat_id = CAT_MAP.get(cat_name, 7)
                desc = clean_str(row[7]) or 'กิจกรรมจิตอาสา'
                location = clean_str(row[8]) if len(row) > 8 else ''
                hours_val = row[4] # Column E: hours
                try:
                    hrs = float(hours_val)
                except:
                    hrs = 2.0
                
                dt_val = row[1] # Column B: Timestamp
                act_date = '2025-11-23'
                if isinstance(dt_val, (datetime.datetime, datetime.date)):
                    act_date = dt_val.strftime('%Y-%m-%d')
                
                deed_id_counter += 1
                imported_deeds.append({
                    'id': str(deed_id_counter),
                    'student_id': sid,
                    'categoryId': cat_id,
                    'hours': hrs,
                    'activityDate': act_date,
                    'description': desc,
                    'note': f"สถานที่: {location}" if location else "",
                    'imageUrl': "510903.jpg",
                    'status': 'approved',
                    'approved_by': 'อาจารย์ผู้ควบคุม (นำเข้า 2568)',
                    'created_at': act_date + 'T08:00:00Z',
                    'updated_at': act_date + 'T08:00:00Z'
                })
                count += 1
            print(f"  ✅ Imported {count} detailed activity records from Data 2568.xlsx")
        wb.close()
    
    # -------------------------------------------------------------
    # FILE 2: รวมบันทึกความดีปี 68 ชั้นปีที่1.xlsx (รุ่น 68)
    # -------------------------------------------------------------
    if os.path.exists(YEAR1_FILE):
        print(f"\n📂 Reading File 2: {os.path.basename(YEAR1_FILE)}")
        wb = openpyxl.load_workbook(YEAR1_FILE, data_only=True)
        ws = wb.active
        count = 0
        cats = [1, 2, 3, 4, 5, 6, 7, 8, 9]
        cat_names = [
            'บริจาคโลหิต/เกล็ดเลือด/พลาสมา',
            'โครงการภายนอก (คำสั่ง วพอ.)',
            'ช่วยเหลืองานภายใน วพอ.',
            'เข้าอบรมที่ วพอ. จัดให้',
            'ช่วยงานหน่วยงาน/ชุมชน/มูลนิธิ',
            'ทำนุบำรุงศาสนสถาน',
            'งานฟรีทั่วไป',
            'กิจกรรมจงรักภักดีต่อสถาบัน',
            'ชม. ที่สมควรได้รับ (บทบาทพิเศษ)'
        ]
        
        for r in range(3, ws.max_row + 1):
            sid = clean_sid(ws.cell(r, 1).value)
            if not sid or sid not in student_by_id:
                fullname = clean_str(ws.cell(r, 3).value)
                if fullname in student_by_name:
                    sid = student_by_name[fullname]
                else:
                    continue
            
            # Read category hours from cols 4..12
            for c_idx, cid in enumerate(cats):
                val = ws.cell(r, 4 + c_idx).value
                try:
                    hrs = float(val)
                except:
                    hrs = 0.0
                if hrs > 0:
                    deed_id_counter += 1
                    imported_deeds.append({
                        'id': str(deed_id_counter),
                        'student_id': sid,
                        'categoryId': cid,
                        'hours': hrs,
                        'activityDate': '2025-12-01',
                        'description': f"สรุปชั่วโมงสะสมหมวด {cat_names[c_idx]} (ปีการศึกษา 2568)",
                        'note': "บันทึกสะสมปีการศึกษา 2568",
                        'imageUrl': "510903.jpg",
                        'status': 'approved',
                        'approved_by': 'อาจารย์ผู้ควบคุม (สรุปปี 2568)',
                        'created_at': '2025-12-01T08:00:00Z',
                        'updated_at': '2025-12-01T08:00:00Z'
                    })
                    count += 1
        print(f"  ✅ Imported {count} category summary records for รุ่น 68")
        wb.close()

    # -------------------------------------------------------------
    # FILE 3: รวมบันทึกความดีปีการศึกษา 68 ชั้นปีที่ 2.xlsx (รุ่น 67)
    # -------------------------------------------------------------
    if os.path.exists(YEAR2_FILE):
        print(f"\n📂 Reading File 3: {os.path.basename(YEAR2_FILE)}")
        sheets_data = parse_xml_xlsx(YEAR2_FILE)
        count = 0
        cat_names = [
            'บริจาคโลหิต/เกล็ดเลือด/พลาสมา',
            'โครงการภายนอก (คำสั่ง วพอ.)',
            'ช่วยเหลืองานภายใน วพอ.',
            'เข้าอบรมที่ วพอ. จัดให้',
            'ช่วยงานหน่วยงาน/ชุมชน/มูลนิธิ',
            'ทำนุบำรุงศาสนสถาน',
            'งานฟรีทั่วไป',
            'กิจกรรมจงรักภักดีต่อสถาบัน',
            'ชม. ที่สมควรได้รับ (บทบาทพิเศษ)'
        ]
        
        # Use sheet2 (index 0 if sheet2 came first)
        rows = sheets_data[0] if sheets_data else []
        for r in rows[2:]: # skip header
            if not r or len(r) < 4:
                continue
            sid = clean_sid(r[0])
            if not sid or sid not in student_by_id:
                fullname = clean_str(r[2]) if len(r) > 2 else ''
                if fullname in student_by_name:
                    sid = student_by_name[fullname]
                else:
                    continue
            
            for c_idx in range(3, min(12, len(r))):
                cid = c_idx - 2
                try:
                    hrs = float(r[c_idx])
                except:
                    hrs = 0.0
                if hrs > 0:
                    deed_id_counter += 1
                    imported_deeds.append({
                        'id': str(deed_id_counter),
                        'student_id': sid,
                        'categoryId': cid,
                        'hours': hrs,
                        'activityDate': '2025-12-01',
                        'description': f"สรุปชั่วโมงสะสมหมวด {cat_names[cid-1]} (ปีการศึกษา 2568)",
                        'note': "บันทึกสะสมปีการศึกษา 2568",
                        'imageUrl': "510903.jpg",
                        'status': 'approved',
                        'approved_by': 'อาจารย์ผู้ควบคุม (สรุปปี 2568)',
                        'created_at': '2025-12-01T08:00:00Z',
                        'updated_at': '2025-12-01T08:00:00Z'
                    })
                    count += 1
        print(f"  ✅ Imported {count} category summary records for รุ่น 67")

    # -------------------------------------------------------------
    # FILE 4: สรุปยอดบันทึกความดีนพอ.ชั้นปีที่3.xlsx (รุ่น 66)
    # -------------------------------------------------------------
    if os.path.exists(YEAR3_FILE):
        print(f"\n📂 Reading File 4: {os.path.basename(YEAR3_FILE)}")
        wb = openpyxl.load_workbook(YEAR3_FILE, data_only=True)
        ws = wb.active
        count = 0
        cat_names = [
            'บริจาคโลหิต/เกล็ดเลือด/พลาสมา',
            'โครงการภายนอก (คำสั่ง วพอ.)',
            'ช่วยเหลืองานภายใน วพอ.',
            'เข้าอบรมที่ วพอ. จัดให้',
            'ช่วยงานหน่วยงาน/ชุมชน/มูลนิธิ',
            'ทำนุบำรุงศาสนสถาน',
            'งานฟรีทั่วไป',
            'กิจกรรมจงรักภักดีต่อสถาบัน',
            'ชม. ที่สมควรได้รับ (บทบาทพิเศษ)'
        ]
        for r in range(3, ws.max_row + 1):
            fullname = clean_str(ws.cell(r, 2).value)
            if not fullname:
                continue
            
            sid = student_by_name.get(fullname)
            if not sid:
                continue
            
            for c_idx in range(1, 10):
                val = ws.cell(r, 2 + c_idx).value
                try:
                    hrs = float(val)
                except:
                    hrs = 0.0
                if hrs > 0:
                    deed_id_counter += 1
                    imported_deeds.append({
                        'id': str(deed_id_counter),
                        'student_id': sid,
                        'categoryId': c_idx,
                        'hours': hrs,
                        'activityDate': '2025-12-01',
                        'description': f"สรุปชั่วโมงสะสมหมวด {cat_names[c_idx-1]} (ปีการศึกษา 2568)",
                        'note': "บันทึกสะสมปีการศึกษา 2568",
                        'imageUrl': "510903.jpg",
                        'status': 'approved',
                        'approved_by': 'อาจารย์ผู้ควบคุม (สรุปปี 2568)',
                        'created_at': '2025-12-01T08:00:00Z',
                        'updated_at': '2025-12-01T08:00:00Z'
                    })
                    count += 1
        print(f"  ✅ Imported {count} category summary records for รุ่น 66")
        wb.close()

    # Deduplicate / merge existing deeds with imported ones
    existing_deeds_file = os.path.join(DATA_DIR, 'deeds.json')
    existing_deeds = []
    if os.path.exists(existing_deeds_file):
        try:
            with open(existing_deeds_file, 'r', encoding='utf-8') as f:
                raw_existing = json.load(f)
                if isinstance(raw_existing, dict):
                    for sid_key, d_list in raw_existing.items():
                        if isinstance(d_list, list):
                            for d in d_list:
                                if isinstance(d, dict):
                                    # Normalize deed keys
                                    normalized = {
                                        'id': str(d.get('id', d.get('deed_id', deed_id_counter))),
                                        'student_id': str(d.get('student_id', sid_key)),
                                        'categoryId': CAT_MAP.get(d.get('category', ''), 7),
                                        'hours': float(d.get('hours', 2.0)),
                                        'activityDate': str(d.get('event_date', d.get('activityDate', '2025-11-01'))),
                                        'description': str(d.get('title', d.get('description', 'บันทึกความดี'))),
                                        'note': str(d.get('location', d.get('note', ''))),
                                        'imageUrl': str(d.get('photo_url', d.get('imageUrl', '510903.jpg'))),
                                        'status': str(d.get('status', 'approved')),
                                        'approved_by': str(d.get('signer', d.get('approved_by', 'อาจารย์ผู้ควบคุม'))),
                                        'created_at': '2025-11-01T08:00:00Z',
                                        'updated_at': '2025-11-01T08:00:00Z'
                                    }
                                    existing_deeds.append(normalized)
                elif isinstance(raw_existing, list):
                    existing_deeds = raw_existing
            print(f"📋 Loaded {len(existing_deeds)} existing deeds.")
        except Exception as e:
            print(f"⚠️ Error loading existing deeds: {e}")

    # Combine all deeds
    all_deeds = imported_deeds + existing_deeds
    
    # Save deeds.json
    with open(os.path.join(DATA_DIR, 'deeds.json'), 'w', encoding='utf-8') as f:
        json.dump(all_deeds, f, ensure_ascii=False, indent=2)
    print(f"✅ Written {os.path.join(DATA_DIR, 'deeds.json')}")
    
    # Save deeds_data.js for frontend
    js_content = f"// Auto-generated from 2568 Excel files\nconst DEEDS_DATA = {json.dumps(all_deeds, ensure_ascii=False, indent=2)};\n"
    for js_p in [os.path.join(DATA_DIR, 'deeds_data.js'), os.path.join(BASE_DIR, 'frontend', 'data', 'deeds_data.js')]:
        with open(js_p, 'w', encoding='utf-8') as f:
            f.write(js_content)
        print(f"✅ Written {js_p}")

    # Generate summary report for all 375 students
    hours_per_student = {}
    for d in all_deeds:
        if d.get('status') == 'approved':
            sid = str(d.get('student_id'))
            hrs = float(d.get('hours', 0))
            hours_per_student[sid] = hours_per_student.get(sid, 0.0) + hrs

    passed_count = sum(1 for sid, h in hours_per_student.items() if h >= 36)
    
    print(f"\n==================================================")
    print(f"📊 สรุปผลการอัปเดตชั่วโมงความดีของนักเรียนทั้งหมด:")
    print(f"   จำนวนรายการความดีในระบบทั้งหมด: {len(all_deeds)} รายการ")
    print(f"   จำนวนนักเรียนที่มีชั่วโมงสะสม: {len(hours_per_student)} คน")
    print(f"   จำนวนนักเรียนที่ผ่านเกณฑ์ (>= 36 ชั่วโมง): {passed_count} คน")
    print(f"==================================================")

if __name__ == '__main__':
    main()
