#!/usr/bin/env python3
"""Export student data from Excel to JSON for import into Google Sheets
   ปีการศึกษา 2569 — อัพเดท 9 มี.ค. 2569
"""
import openpyxl
import json
import re
import os

DATA_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(DATA_DIR)
EXCEL_FILE = os.path.join(BASE_DIR, "รายชื่อ นพอ.ปี68 ทุกชั้นปี ( 5 ก.พ. 69 ตัดเกาหลี )SWD .xlsx")

# Sheet -> class year mapping (ปีการศึกษา 2569)
SHEET_CLASS_MAP = {
    'SWD69': {'year': 1, 'class_year': 69},  # ชั้นปีที่ 1 (รุ่น 69) - ยังไม่มีในไฟล์
    'SWD68': {'year': 2, 'class_year': 68},  # ชั้นปีที่ 2 (รุ่น 68)
    'SWD67': {'year': 3, 'class_year': 67},  # ชั้นปีที่ 3 (รุ่น 67)
    'SWD66': {'year': 4, 'class_year': 66},  # ชั้นปีที่ 4 (รุ่น 66)
    'SWD65': {'year': 5, 'class_year': 65},  # ศิษย์เก่า (รุ่น 65)
    'SWD64': {'year': 5, 'class_year': 64},  # ศิษย์เก่า (รุ่น 64)
}

def clean_number(val):
    if val is None:
        return None
    s = str(val).strip().replace('*', '')
    try:
        return int(float(s))
    except:
        return None

def main():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    students = []
    
    # Load existing student profiles to merge edits (prevent data loss)
    existing_students = {}
    frontend_json_path = os.path.join(BASE_DIR, 'frontend', 'data', 'students.json')
    if os.path.exists(frontend_json_path):
        try:
            with open(frontend_json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                existing_students = {s['student_id']: s for s in data}
            print(f"Loaded {len(existing_students)} existing student records to preserve edits/passwords.")
        except Exception as e:
            print(f"⚠️ Failed to load existing students.json for merging: {e}")

    for sheet_name, meta in SHEET_CLASS_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        
        for row in ws.iter_rows(min_row=5, values_only=True):
            student_id = clean_number(row[1])
            if student_id is None:
                continue
            
            rank = str(row[2]).strip() if row[2] else 'นพอ.'
            first_name = str(row[3]).strip() if row[3] else ''
            last_name = str(row[4]).strip() if row[4] else ''
            note = str(row[5]).strip() if row[5] else ''
            
            student_id_str = str(student_id)
            password = student_id_str
            email = ''
            telegram_chat_id = ''
            role = 'student'
            class_year = meta['class_year']
            year_level = meta['year']
            
            position = 'นักเรียนพยาบาล'
            nickname = ''
            phone = ''

            # Merge edits from existing database if present
            if student_id_str in existing_students:
                existing = existing_students[student_id_str]
                rank = existing.get('rank', rank)
                first_name = existing.get('first_name', first_name)
                last_name = existing.get('last_name', last_name)
                password = existing.get('password', password)
                email = existing.get('email', email)
                telegram_chat_id = existing.get('telegram_chat_id', telegram_chat_id)
                role = existing.get('role', role)
                class_year = existing.get('class_year', class_year)
                year_level = existing.get('year_level', year_level)
                note = existing.get('note', note)
                position = existing.get('position', position)
                nickname = existing.get('nickname', nickname)
                phone = existing.get('phone', phone)

            student = {
                'student_id': student_id_str,
                'rank': rank,
                'first_name': first_name,
                'last_name': last_name,
                'full_name': f"{first_name} {last_name}".strip(),
                'nickname': nickname,
                'phone': phone,
                'class_year': class_year,
                'year_level': year_level,
                'note': note,
                'position': position,
                'password': password,
                'email': email,
                'telegram_chat_id': telegram_chat_id,
                'role': role
            }
            students.append(student)

    # Append Class 69 test students since they are not in the Excel sheets yet
    class69_students = [
        {
            "student_id": "6900001",
            "rank": "นพอ.",
            "first_name": "กิตติภพ",
            "last_name": "ทองดี",
            "full_name": "กิตติภพ ทองดี",
            "class_year": 69,
            "year_level": 1,
            "note": "นักเรียนใหม่ชั้นปีที่ 1",
            "password": "6900001",
            "email": "",
            "telegram_chat_id": "",
            "role": "student"
        },
        {
            "student_id": "6900002",
            "rank": "นพอ.",
            "first_name": "พรนภัส",
            "last_name": "จิตใจดี",
            "full_name": "พรนภัส จิตใจดี",
            "class_year": 69,
            "year_level": 1,
            "note": "นักเรียนใหม่ชั้นปีที่ 1",
            "password": "6900002",
            "email": "",
            "telegram_chat_id": "",
            "role": "student"
        },
        {
            "student_id": "6900003",
            "rank": "นพอ.",
            "first_name": "วรเมธ",
            "last_name": "รักสงบ",
            "full_name": "วรเมธ รักสงบ",
            "class_year": 69,
            "year_level": 1,
            "note": "นักเรียนใหม่ชั้นปีที่ 1",
            "password": "6900003",
            "email": "",
            "telegram_chat_id": "",
            "role": "student"
        },
        {
            "student_id": "6900004",
            "rank": "นพอ.",
            "first_name": "ชนม์นิภา",
            "last_name": "มีสุข",
            "full_name": "ชนม์นิภา มีสุข",
            "class_year": 69,
            "year_level": 1,
            "note": "นักเรียนใหม่ชั้นปีที่ 1",
            "password": "6900004",
            "email": "",
            "telegram_chat_id": "",
            "role": "student"
        },
        {
            "student_id": "6900005",
            "rank": "นพอ.",
            "first_name": "ปองพล",
            "last_name": "คนดี",
            "full_name": "ปองพล คนดี",
            "class_year": 69,
            "year_level": 1,
            "note": "นักเรียนใหม่ชั้นปีที่ 1",
            "password": "6900005",
            "email": "",
            "telegram_chat_id": "",
            "role": "student"
        }
    ]
    
    # Merge existing class69_students modifications if they exist in the DB
    for s in class69_students:
        sid = s['student_id']
        if sid in existing_students:
            existing = existing_students[sid]
            s['rank'] = existing.get('rank', s['rank'])
            s['first_name'] = existing.get('first_name', s['first_name'])
            s['last_name'] = existing.get('last_name', s['last_name'])
            s['full_name'] = existing.get('full_name', s['full_name'])
            s['class_year'] = existing.get('class_year', s['class_year'])
            s['year_level'] = existing.get('year_level', s['year_level'])
            s['password'] = existing.get('password', s['password'])
            s['email'] = existing.get('email', s['email'])
            s['telegram_chat_id'] = existing.get('telegram_chat_id', s['telegram_chat_id'])
            s['role'] = existing.get('role', s['role'])
            s['note'] = existing.get('note', s['note'])

    students.extend(class69_students)
    
    # Missing historical students from Main 2568.xlsx who are not in the new rosters
    missing_historical_students = [
        {
            "student_id": "6503719",
            "rank": "นพอ.",
            "first_name": "นันท์นภัส",
            "last_name": "เภสัชชา",
            "full_name": "นันท์นภัส เภสัชชา",
            "class_year": 65,
            "year_level": 5,
            "note": "ข้อมูลนำเข้าย้อนหลังจากประวัติปี 2568",
            "password": "6503719",
            "email": "",
            "telegram_chat_id": "",
            "role": "student"
        },
        {
            "student_id": "6603775",
            "rank": "นพอ.",
            "first_name": "ธิดารัตน์",
            "last_name": "นิลสังข์",
            "full_name": "ธิดารัตน์ นิลสังข์",
            "class_year": 66,
            "year_level": 4,
            "note": "ข้อมูลนำเข้าย้อนหลังจากประวัติปี 2568",
            "password": "6603775",
            "email": "",
            "telegram_chat_id": "",
            "role": "student"
        },
        {
            "student_id": "6703840",
            "rank": "นพอ.",
            "first_name": "ดลภัค",
            "last_name": "แก้วเอก",
            "full_name": "ดลภัค แก้วเอก",
            "class_year": 67,
            "year_level": 3,
            "note": "ข้อมูลนำเข้าย้อนหลังจากประวัติปี 2568",
            "password": "6703840",
            "email": "",
            "telegram_chat_id": "",
            "role": "student"
        },
        {
            "student_id": "6703850",
            "rank": "นพอ.",
            "first_name": "ประริชญา",
            "last_name": "ประสิทธิ์พรม",
            "full_name": "ประริชญา ประสิทธิ์พรม",
            "class_year": 67,
            "year_level": 3,
            "note": "ข้อมูลนำเข้าย้อนหลังจากประวัติปี 2568",
            "password": "6703850",
            "email": "",
            "telegram_chat_id": "",
            "role": "student"
        },
        {
            "student_id": "6703871",
            "rank": "นพอ.",
            "first_name": "วีรภัทร",
            "last_name": "นกดำ",
            "full_name": "วีรภัทร นกดำ",
            "class_year": 67,
            "year_level": 3,
            "note": "ข้อมูลนำเข้าย้อนหลังจากประวัติปี 2568",
            "password": "6703871",
            "email": "",
            "telegram_chat_id": "",
            "role": "student"
        }
    ]
    
    # Merge existing modifications for missing historical students
    for s in missing_historical_students:
        sid = s['student_id']
        if sid in existing_students:
            existing = existing_students[sid]
            s['rank'] = existing.get('rank', s['rank'])
            s['first_name'] = existing.get('first_name', s['first_name'])
            s['last_name'] = existing.get('last_name', s['last_name'])
            s['full_name'] = existing.get('full_name', s['full_name'])
            s['class_year'] = existing.get('class_year', s['class_year'])
            s['year_level'] = existing.get('year_level', s['year_level'])
            s['password'] = existing.get('password', s['password'])
            s['email'] = existing.get('email', s['email'])
            s['telegram_chat_id'] = existing.get('telegram_chat_id', s['telegram_chat_id'])
            s['role'] = existing.get('role', s['role'])
            s['note'] = existing.get('note', s['note'])

    students.extend(missing_historical_students)
    
    # Sort by class year then student_id
    students.sort(key=lambda x: (x['class_year'], x['student_id']))
    
    print(f"Total students exported: {len(students)}")
    
    # Print breakdown
    from collections import Counter
    years = Counter(s['class_year'] for s in students)
    for y, count in sorted(years.items()):
        print(f"  Class {y}: {count} students")
    
    # Save to JSON
    json_path = os.path.join(DATA_DIR, 'students.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(students, f, ensure_ascii=False, indent=2)
    
    print(f"\nSaved to {json_path}")
    
    # Also save as JS module for frontend use
    js_path = os.path.join(DATA_DIR, 'students_data.js')
    with open(js_path, 'w', encoding='utf-8') as f:
        f.write("// Auto-generated student data - DO NOT EDIT MANUALLY\n")
        f.write("// Generated from: รายชื่อ นพอ.ปี69 ทุกชั้นปี\n\n")
        f.write("const STUDENTS_DATA = ")
        json.dump(students, f, ensure_ascii=False, indent=2)
        f.write(";\n\n")
        f.write("if (typeof window !== 'undefined') { window.STUDENTS_DATA = STUDENTS_DATA; }\n")
        f.write("if (typeof globalThis !== 'undefined') { globalThis.STUDENTS_DATA = STUDENTS_DATA; }\n")
    
    print(f"Saved to {js_path}")

    # Write to frontend/data/
    frontend_json_path = os.path.join(BASE_DIR, 'frontend', 'data', 'students.json')
    frontend_js_path = os.path.join(BASE_DIR, 'frontend', 'data', 'students_data.js')
    os.makedirs(os.path.dirname(frontend_js_path), exist_ok=True)
    
    with open(frontend_json_path, 'w', encoding='utf-8') as f:
        json.dump(students, f, ensure_ascii=False, indent=2)
    
    with open(frontend_js_path, 'w', encoding='utf-8') as f:
        f.write("// Auto-generated student data - DO NOT EDIT MANUALLY\n")
        f.write("// Generated from: รายชื่อ นพอ.ปี69 ทุกชั้นปี\n\n")
        f.write("const STUDENTS_DATA = ")
        json.dump(students, f, ensure_ascii=False, indent=2)
        f.write(";\n\n")
        f.write("if (typeof window !== 'undefined') { window.STUDENTS_DATA = STUDENTS_DATA; }\n")
        f.write("if (typeof globalThis !== 'undefined') { globalThis.STUDENTS_DATA = STUDENTS_DATA; }\n")
        
    print(f"Sync-copied to {frontend_js_path}")

if __name__ == '__main__':
    main()

