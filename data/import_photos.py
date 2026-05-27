#!/usr/bin/env python3
"""
import_photos.py
จับคู่รูปภาพกับนักเรียนจาก Excel + โฟลเดอร์รูป
แล้วสร้าง students_photos.js สำหรับใช้ใน frontend

วิธีใช้:
    python3 import_photos.py

ผลลัพธ์:
    data/students_photos.js  — const STUDENT_PHOTOS = { "student_id": "data:image/jpeg;base64,..." }
    data/students_data.js    — อัพเดทรายชื่อนักเรียนจาก Excel
"""
import openpyxl, json, base64, os, re
from pathlib import Path
from datetime import datetime

BASE_DIR   = Path(__file__).parent.parent  # งานบันทึกความดี 69/
DATA_DIR   = Path(__file__).parent
EXCEL_FILE = BASE_DIR / "รายชื่อ นพอ.ปี68 ทุกชั้นปี ( 5 ก.พ. 69 ตัดเกาหลี )SWD .xlsx"

# โฟลเดอร์รูปแต่ละชั้นปี (ปีการศึกษา 2569) ในโฟลเดอร์จัดเก็บแบบมืออาชีพ data/raw_photos/
RAW_PHOTOS_DIR = DATA_DIR / "raw_photos"
PHOTO_DIRS = {
    68: RAW_PHOTOS_DIR / "นพอ.ชั้นปีที่ 1 ",   # SWD68 = ชั้นปีที่ 2 (รุ่น 68)
    67: RAW_PHOTOS_DIR / "ชั้นปีที่ 2",          # SWD67 = ชั้นปีที่ 3 (รุ่น 67)
    66: RAW_PHOTOS_DIR / "ชั้นปีที่ 3",          # SWD66 = ชั้นปีที่ 4 (รุ่น 66)
    65: RAW_PHOTOS_DIR / "ชั้นปีที่ 4",          # SWD65 = ศิษย์เก่า (รุ่น 65)
}

OUTPUT_JS    = DATA_DIR / "students_photos.js"
STUDENTS_JS  = DATA_DIR / "students_data.js"

SHEET_CLASS_MAP = {
    "SWD68": {"year": 2, "class_year": 68},  # ชั้นปีที่ 2 (ปีการศึกษา 2569)
    "SWD67": {"year": 3, "class_year": 67},  # ชั้นปีที่ 3
    "SWD66": {"year": 4, "class_year": 66},  # ชั้นปีที่ 4
    "SWD65": {"year": 5, "class_year": 65},  # ศิษย์เก่า รุ่น 65
    "SWD64": {"year": 5, "class_year": 64},  # ศิษย์เก่า รุ่น 64 (ไม่มีรูป)
}


# -------- helpers --------
def clean_id(val):
    if val is None: return None
    s = str(val).strip().replace("*", "").replace(" ", "")
    try: return str(int(float(s)))
    except: return None

def encode_photo(path: Path) -> str:
    ext = path.suffix.lower().lstrip(".")
    mime = {"jpg": "jpeg", "jpeg": "jpeg", "png": "png", "gif": "gif",
            "webp": "webp"}.get(ext, "jpeg")
    data = base64.b64encode(path.read_bytes()).decode()
    return f"data:image/{mime};base64,{data}"

def get_sorted_photos(folder: Path):
    """ดึงไฟล์รูปเรียงตามหมายเลขนำหน้า"""
    if not folder.exists():
        print(f"⚠️  ไม่พบโฟลเดอร์: {folder}")
        return []
    imgs = []
    for f in folder.iterdir():
        if f.suffix.lower() in (".jpg", ".jpeg", ".png", ".gif", ".webp") and not f.name.startswith("."):
            m = re.match(r"^(\d+)", f.name.strip())
            num = int(m.group(1)) if m else 9999
            imgs.append((num, f))
    imgs.sort(key=lambda x: x[0])
    return imgs

# -------- read Excel --------
def read_students(wb):
    students = []
    for sheet_name, meta in SHEET_CLASS_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"⚠️  ไม่พบ sheet: {sheet_name}")
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=5, values_only=True))  # row 5+ = data
        row_num = 0
        for row in rows:
            sid = clean_id(row[1])  # col B
            if sid is None: continue
            # col C = ยศ, col D = ชื่อ, col E = สกุล
            rank  = str(row[2]).strip() if row[2] else "นพอ."
            fname = str(row[3]).strip() if row[3] else ""
            lname = str(row[4]).strip() if row[4] else ""
            note  = str(row[5]).strip() if row[5] else ""
            row_num += 1
            students.append({
                "student_id":  sid,
                "rank":        rank,
                "first_name":  fname,
                "last_name":   lname,
                "full_name":   f"{fname} {lname}".strip(),
                "class_year":  meta["class_year"],
                "year_level":  meta["year"],
                "note":        note,
                "password":    sid,
                "email":       "",
                "telegram_chat_id": "",
                "role":        "student",
                "_sheet":      sheet_name,
                "_row_num":    row_num,   # 1-based index ใน sheet
            })
    students.sort(key=lambda x: (x["class_year"], x["student_id"]))
    return students

# -------- match photos --------
def extract_name_from_filename(img_path: Path) -> str:
    """ดึงชื่อ-สกุลจากชื่อไฟล์รูป เช่น '01_นพอ.กมลฉัตร ชาสุรีย์.jpg' → 'กมลฉัตร ชาสุรีย์'"""
    stem = img_path.stem.strip()
    # ตัดหมายเลขนำหน้า เช่น "01_", "01 "
    stem = re.sub(r"^\d+[_\s]+", "", stem)
    # ตัดคำนำหน้า เช่น "นพอ.(ช) ", "นพอ. "
    stem = re.sub(r"^นพอ\.\s*(\(ช\))?\s*", "", stem).strip()
    return stem

def match_photos(students):
    photos = {}  # student_id → base64 string
    for class_year, folder in PHOTO_DIRS.items():
        sorted_imgs = get_sorted_photos(folder)
        if not sorted_imgs:
            continue
        # นักเรียนในชั้นปีนี้ เรียงตาม _row_num
        class_students = sorted(
            [s for s in students if s["class_year"] == class_year],
            key=lambda x: x["_row_num"]
        )
        # สร้าง lookup ตามชื่อ-สกุล
        name_to_stu = {f'{s["first_name"]} {s["last_name"]}'.strip(): s
                       for s in class_students}

        matched = 0
        unmatched = []
        for img_num, img_path in sorted_imgs:
            stu = None
            name_in_file = extract_name_from_filename(img_path)
            # 1) จับคู่จากชื่อในชื่อไฟล์เป็นหลัก
            stu = name_to_stu.get(name_in_file)
            if stu:
                print(f"    🔍 จับคู่ชื่อ: [{img_path.name}] → {stu['student_id']} {name_in_file}")
            
            # 2) ถ้าจับคู่ชื่อไม่ได้ ลองจับคู่จากหมายเลข (_row_num)
            if stu is None:
                stu = next((s for s in class_students if s["_row_num"] == img_num), None)
                if stu:
                    print(f"    ⚠️ จับคู่ด้วยหมายเลข: [{img_path.name}] → {stu['student_id']} (ชื่ออาจไม่ตรง)")
            if stu:
                try:
                    photos[stu["student_id"]] = encode_photo(img_path)
                    matched += 1
                except Exception as e:
                    print(f"  ⚠️  {img_path.name}: {e}")
            else:
                unmatched.append(img_path.name)
        print(f"  📷 รุ่น {class_year}: จับคู่รูปได้ {matched}/{len(class_students)} คน "
              f"(มีรูป {len(sorted_imgs)} ไฟล์)")
        if unmatched:
            for u in unmatched:
                print(f"    ⚠️  ไม่พบนักเรียนสำหรับรูป: {u}")
    return photos


# -------- write outputs --------
def write_photos_js(photos):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    photo_js = (
        f"// Auto-generated student photos — {ts}\n"
        f"// นักเรียนที่มีรูป: {len(photos)} คน  DO NOT EDIT MANUALLY\n"
        f"const STUDENT_PHOTOS = {json.dumps(photos, ensure_ascii=False)};\n"
    )
    OUTPUT_JS.write_text(photo_js, encoding="utf-8")
    
    # Write to frontend/data/
    frontend_output_js = BASE_DIR / "frontend" / "data" / "students_photos.js"
    frontend_output_js.parent.mkdir(parents=True, exist_ok=True)
    frontend_output_js.write_text(photo_js, encoding="utf-8")
    
    print(f"\n✅ บันทึกรูปภาพ {len(photos)} คน → {OUTPUT_JS}")
    print(f"   Sync-copied to {frontend_output_js}")

def write_students_js(students):
    clean = [{k: v for k, v in s.items() if not k.startswith("_")} for s in students]
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    js = (
        f"// Auto-generated student data — {ts}  DO NOT EDIT MANUALLY\n"
        f"// นักเรียนทั้งหมด: {len(clean)} คน\n"
        "const STUDENTS_DATA = "
        + json.dumps(clean, ensure_ascii=False, indent=2) + ";\n"
    )
    STUDENTS_JS.write_text(js, encoding="utf-8")
    
    # Write to frontend/data/
    frontend_students_js = BASE_DIR / "frontend" / "data" / "students_data.js"
    frontend_students_js.write_text(js, encoding="utf-8")
    
    print(f"✅ บันทึกรายชื่อ {len(clean)} คน → {STUDENTS_JS}")
    print(f"   Sync-copied to {frontend_students_js}")


# -------- main --------
def main():
    print("📂 อ่านข้อมูลจาก Excel...")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    students = read_students(wb)
    print(f"   นักเรียนทั้งหมด: {len(students)} คน")

    print("\n🖼️  จับคู่รูปภาพ...")
    photos = match_photos(students)

    write_photos_js(photos)
    write_students_js(students)

    # สรุปตามชั้นปี
    from collections import Counter
    year_count = Counter(s["class_year"] for s in students)
    print("\n📊 สรุปตามชั้นปี:")
    labels = {69:"ชั้นปีที่ 1", 68:"ชั้นปีที่ 2", 67:"ชั้นปีที่ 3",
              66:"ชั้นปีที่ 4", 65:"ศิษย์เก่า (65)", 64:"ศิษย์เก่า (64)"}
    for cy, cnt in sorted(year_count.items(), reverse=True):
        print(f"   รุ่น {cy} ({labels.get(cy,'?')}): {cnt} คน")

if __name__ == "__main__":
    main()
