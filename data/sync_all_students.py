#!/usr/bin/env python3
"""Build a PRIVATE server-side roster only.

Security invariants:
- Never export student roster, names, passwords, LINE IDs, Telegram credentials,
  or application secrets into frontend/, public/, docs/, or tracked repository data.
- Output is written only to private-data/, which must stay gitignored.
- Frontend authentication must use a trusted server-side endpoint.
"""
import json
import os
import re
import openpyxl

DATA_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(DATA_DIR)
PRIVATE_DIR = os.path.join(BASE_DIR, "private-data")
PRIVATE_ROSTER = os.path.join(PRIVATE_DIR, "students.json")

EXCEL_FILES = [
    os.path.join(DATA_DIR, "รายชื่อ_นพอ_ปี69_2569.xlsx"),
    os.path.join(DATA_DIR, "รายชื่อ_นพอ_อัปเดตล่าสุด_2569.xlsx"),
    os.path.join(BASE_DIR, "รายชื่อ นพอ.ปี 69.xlsx"),
    os.path.join(BASE_DIR, "รายชื่อ นพอ.ปี68 ทุกชั้นปี ( 5 ก.พ. 69 ตัดเกาหลี )SWD .xlsx"),
]

SHEET_MAP = {
    "SWD69": 69, "SWD68": 68, "SWD67": 67, "SWD66": 66,
    "SWD65": 65, "SWD64": 64,
    "นพอ.1": 68, "นพอ.1 ": 68, "นพอ.2": 67, "นพอ.3": 66,
    "นพอ.4": 65, "นพอ.5": 65,
    "นพอ.ปี1": 69, "นพอ.ปี2": 68, "นพอ.ปี3": 67, "นพอ.ปี4": 66,
}


def get_study_year(class_year):
    return {69: 1, 68: 2, 67: 3, 66: 4, 65: 5, 64: 5}.get(class_year, 5)


def clean_number(value):
    if value is None:
        return None
    text = str(value).strip().replace("*", "").replace(" ", "")
    try:
        return str(int(float(text)))
    except (TypeError, ValueError):
        return None


def clean_rank(value):
    text = re.sub(r"\s+", " ", str(value or "").strip())
    return text or "นพอ."


def extract_students_from_sheet(ws, class_year):
    students = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row or len(row) < 5:
            continue
        student_id = clean_number(row[1])
        if not student_id or len(student_id) < 6:
            continue
        first_name = str(row[3] or "").strip()
        if not first_name:
            continue
        students.append({
            "student_id": student_id,
            "rank": clean_rank(row[2]),
            "first_name": first_name,
            "last_name": str(row[4] or "").strip(),
            "class_year": class_year,
            "year": get_study_year(class_year),
        })
    return students


def assert_private_output(path):
    resolved = os.path.realpath(path)
    private_root = os.path.realpath(PRIVATE_DIR) + os.sep
    if not resolved.startswith(private_root):
        raise RuntimeError("Security refusal: roster output must remain in private-data/")


def main():
    all_students = {}
    for excel_path in EXCEL_FILES:
        if not os.path.exists(excel_path):
            continue
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
        try:
            for sheet_name in wb.sheetnames:
                if sheet_name in ("Settings", "ตั้งค่าระบบ"):
                    continue
                class_year = SHEET_MAP.get(sheet_name.strip()) or SHEET_MAP.get(sheet_name)
                if class_year is None:
                    continue
                for student in extract_students_from_sheet(wb[sheet_name], class_year):
                    all_students.setdefault(student["student_id"], student)
        finally:
            wb.close()

    os.makedirs(PRIVATE_DIR, mode=0o700, exist_ok=True)
    assert_private_output(PRIVATE_ROSTER)
    with open(PRIVATE_ROSTER, "w", encoding="utf-8") as handle:
        json.dump(sorted(all_students.values(), key=lambda x: x["student_id"]), handle,
                  ensure_ascii=False, indent=2)
    try:
        os.chmod(PRIVATE_ROSTER, 0o600)
    except OSError:
        pass
    print(f"Private roster prepared: {len(all_students)} records. No frontend export performed.")


if __name__ == "__main__":
    main()
